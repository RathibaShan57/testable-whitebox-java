#!/usr/bin/env python3
"""Cross-reference taxonomy gate HTML with downloaded tool output artifacts."""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Import parsers from existing script
sys.path.insert(0, str(Path(__file__).resolve().parent))
from parse_taxonomy_gate_report import (  # noqa: E402
    RUNNER_TOOL_MAP,
    TECHNIQUE_TOOL_MAP,
    build_tool_summary,
    parse_report,
)

TOOL_DIR = Path(
    r"c:\Users\Rathiba\OneDrive - testable.cloud\Testable - Execution team - QA-Testable"
    r"\QA Daily Status\QA TestCases\Java\whitebox_java_repo"
)
REPORT_HTML = Path(
    r"c:\Users\Rathiba\Downloads\taxonomy-gate-09998622-2a95-44d0-85f8-910e062f59f5.html"
)
OUT_XLSX = Path(
    r"c:\Users\Rathiba\Downloads\taxonomy-gate-09998622-tool-output-analysis.xlsx"
)

# Expected Java primary tools for testable-whitebox-java (from repo README / mapping)
JAVA_EXPECTED_TOOLS = [
    ("CK", "class.csv", "White Box cyclomatic / coupling metrics"),
    ("PMD (Checkstyle+PMD runner)", "pmd*.xml", "Lint, cognitive complexity, cyclomatic"),
    ("PMD (java-perf-dependency)", "target/pmd.xml or perf-dependency-java*", "Perf dependency analysis — gate expects target/pmd.xml"),
    ("CPD", "cpd*.xml", "Code duplication"),
    ("Checkstyle", "checkstyle*.xml", "Lint / rule violations"),
    ("SpotBugs + FindSecBugs", "spotbugs*.xml", "SAST / security"),
    ("OWASP Dependency-Check", "dependency-check-report*.xml", "SCA / CVE"),
    ("JaCoCo (current)", "jacoco*.xml or *.crdownload JaCoCo XML", "Statement/branch/path coverage"),
    ("JaCoCo (baseline)", "jacoco-baseline.xml", "Coverage delta baseline"),
    ("JaCoCo Coverage Delta", "coverage_delta*.json", "Delta between current and baseline"),
    ("PIT", "pit*.xml or mutations.xml", "Mutation testing"),
    ("Git churn / pydriller", "git-churn*.json", "Code churn / technical debt"),
    ("Lizard", "lizard*.xml", "Performance complexity"),
    ("Semgrep", "semgrep*.json", "Perf AST, PII, security patterns"),
    ("Gitleaks", "gitleaks*.json", "Hardcoded secrets"),
    ("detect-secrets", "detect_secrets*.json", "Secret baseline scan"),
    ("Trufflehog (filesystem)", "trufflehog*.json (non-git)", "Filesystem secrets"),
    ("Trufflehog (git history)", "trufflehog*git*.json", "Historical secrets"),
    ("Checkov", "checkov*.json", "IaC policy"),
    ("tfsec", "tfsec*.json", "Terraform security"),
    ("kics", "kics*", "IaC scanning"),
    ("Microsoft Presidio", "presidio*.json", "PII in fixtures"),
    ("GitHub API (collaborators)", "collaborators.json or access_control*.json", "Privileged access audit"),
    ("GitHub Branch Protection", "github_branch_protection*.json", "Change management"),
    ("ArchUnit", "surefire / archunit*", "Circular dependency (test phase)"),
    ("Newman / Postman", "newman*.json", "Black box API"),
    ("k6", "k6*.json or summary.json", "Load testing"),
    ("Playwright", "playwright*", "UI black box"),
    ("OWASP ZAP", "zap*.json", "DAST"),
]

# Artifact files observed in QA folder → interpretation
ARTIFACT_NOTES = {
    "perf-dependency-python.json": "Python pylint runner output — NOT Java java-perf-dependency PMD",
    "perf-dependency-javascript.json": "JavaScript dep scan — not applicable to Java repo",
    "defs_uses.json": "Beniget scanned Python scripts in sandbox, not Java sources",
    "def_use (1).json": "Python def-use stub",
    "js_all_defs_uses.json": "JavaScript def-use — N/A for Java",
    "js_path_coverage.json": "JavaScript path coverage — N/A for Java",
    "coverage (1).json": "Python coverage.py stub (53 bytes)",
    "coverage (2).json": "Python coverage.py stub (53 bytes)",
    "cobertura-coverage.xml": "JS/Python cobertura proxy — not JaCoCo",
    "coverage-summary.json": "Likely JS/Istanbul summary — not JaCoCo",
    "pylint (1).json": "Python pylint — N/A for Java whitebox metrics",
    "pip_audit.json": "Python pip-audit — N/A for Java Maven repo",
    "bandit (1).json": "Python Bandit — N/A for Java",
    "radon_cc.json": "Python Radon — N/A (Java uses CK+PMD)",
    "npm-audit.json": "Node/npm — blackbox folder only",
    "eslint.json": "JavaScript ESLint — N/A",
    "jscpd-report (1).json": "JS clone detect — Java uses CPD",
    "madge-circular.json": "JS circular deps — Java uses ArchUnit",
    "depcruise.json": "JS dependency-cruiser empty",
    "cosmic_ray.json": "Python mutation — skipped (no Python module)",
    "testmon.json": "Python testmon — N/A",
    "crosshair.json": "Python symbolic execution — N/A",
    "pymcdc.json": "Python MC/DC — N/A",
    "cognitive_ast.json": "Python cognitive AST — Java uses PMD",
    "pydriller_skip (2).json": "Pydriller skipped — no_git_repository in that sandbox",
    "access_control (1).json": "GitHub API skipped: not_github_or_gitlab",
    "coverage_delta (1).json": "JS coverage delta — baseline_missing",
    "coverage_delta (2).json": "Coverage delta all zeros — baseline/current not resolved",
}


def glob_matches(folder: Path, pattern: str) -> list[Path]:
    if " or " in pattern:
        found: list[Path] = []
        for part in pattern.split(" or "):
            found.extend(glob_matches(folder, part.strip()))
        return sorted(set(found))
    if pattern.endswith("*"):
        return sorted(folder.glob(pattern))
    if "*" in pattern:
        return sorted(folder.glob(pattern))
    p = folder / pattern
    return [p] if p.exists() else []


def classify_artifact(path: Path) -> str:
    name = path.name.lower()
    if "jacoco" in name or (path.suffix == ".crdownload" and path.read_bytes()[:50].find(b"JACOCO") >= 0):
        return "JaCoCo XML"
    if "pmd" in name and path.suffix == ".xml":
        return "PMD XML"
    if "dependency-check" in name:
        return "OWASP Dependency-Check"
    if "gitleaks" in name:
        return "Gitleaks"
    if "git-churn" in name or "git_churn" in name:
        return "Git churn"
    if "coverage_delta" in name:
        return "Coverage Delta JSON"
    if "checkstyle" in name:
        return "Checkstyle"
    if "cpd" in name:
        return "CPD"
    if "class.csv" in name:
        return "CK"
    if "semgrep" in name:
        return "Semgrep"
    if "checkov" in name:
        return "Checkov"
    if "tfsec" in name:
        return "tfsec"
    if "presidio" in name:
        return "Presidio"
    if "detect_secrets" in name:
        return "detect-secrets"
    if "lizard" in name:
        return "Lizard"
    if "access_control" in name or "collaborators" in name:
        return "GitHub API"
    if "github_branch" in name:
        return "GitHub Branch Protection"
    return "Other / non-Java"


def artifact_quality(path: Path) -> tuple[str, str]:
    """Return (status, detail) for artifact usability."""
    name = path.name
    note = ARTIFACT_NOTES.get(name, "")
    try:
        size = path.stat().st_size
    except OSError:
        return "Missing", "File not found"

    if size <= 3 and path.suffix == ".json":
        return "Empty/stub", f"{size} bytes — likely skipped or no findings ({note or 'minimal JSON'})"

    if name in ARTIFACT_NOTES:
        return "Wrong context", ARTIFACT_NOTES[name]

    if name.startswith("coverage_delta"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if data.get("skip_reason") == "baseline_missing":
                return "Invalid", "baseline_missing — no delta computable"
            if data.get("current_total") == 0 and data.get("baseline_total") == 0:
                return "Invalid", "current_total and baseline_total are 0 — gate cannot compute delta"
        except json.JSONDecodeError:
            return "Invalid", "Malformed JSON"

    if name == "access_control (1).json":
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if data.get("skipped"):
                return "Skipped", data.get("reason", "skipped")
        except json.JSONDecodeError:
            pass

    if name == "pydriller_skip (2).json":
        return "Skipped", "no_git_repository in worker sandbox"

    if path.suffix == ".crdownload":
        head = path.read_bytes()[:200]
        if b"JACOCO" in head or b"jacoco" in head.lower():
            return "Valid", f"JaCoCo XML ({size:,} bytes)"
        if b"pmd" in head.lower():
            return "Valid", f"PMD XML ({size:,} bytes)"
        return "Unknown", f"Download incomplete? ({size:,} bytes)"

    if size > 100:
        return "Valid", f"{size:,} bytes"
    return "Minimal", f"{size:,} bytes"


def map_expected_coverage() -> list[dict]:
    rows = []
    all_files = list(TOOL_DIR.iterdir()) if TOOL_DIR.is_dir() else []
    names = {p.name for p in all_files}

    for tool, pattern, purpose in JAVA_EXPECTED_TOOLS:
        matches = glob_matches(TOOL_DIR, pattern)
        # Special: JaCoCo in crdownload
        if "jacoco" in pattern.lower() and not matches:
            for p in all_files:
                if p.suffix == ".crdownload":
                    try:
                        if b"JACOCO" in p.read_bytes()[:100]:
                            matches.append(p)
                    except OSError:
                        pass

        if matches:
            statuses = [artifact_quality(p) for p in matches]
            overall = "Present"
            if any(s[0] in ("Invalid", "Skipped", "Wrong context", "Empty/stub") for s in statuses):
                overall = "Present but unusable"
            detail = "; ".join(f"{p.name}: {s[0]} — {s[1]}" for p, s in zip(matches, statuses))
        else:
            overall = "Missing"
            detail = f"No file matching '{pattern}' in QA output folder"

        rows.append(
            {
                "Expected Java Tool": tool,
                "File Pattern": pattern,
                "Purpose": purpose,
                "Coverage Status": overall,
                "Matched Files": ", ".join(p.name for p in matches) if matches else "—",
                "Artifact Detail": detail,
            }
        )
    return rows


def evidence_reason(row: dict, artifact_index: dict) -> str:
    note = row.get("Evidence Note", "")
    tool = row.get("Primary Tool (mapped)", "")
    l3 = row.get("L3 Technique", "")
    exec_status = row.get("Tool Execution Status", "")
    has_paths = bool(row.get("Evidence Paths (sample)", "").strip())

    if "Evidence not available" not in note:
        if "Tool failure" in note:
            return "Tool runner failed — see Evidence Note"
        if "Planned tasks did not start" in note:
            return "Task never scheduled on worker (TOOL_PROFILE / runner wiring)"
        if "Baseline unavailable" in exec_status or "Baseline unavailable" in note:
            return "Coverage delta baseline not resolved (coverage_delta JSON all zeros / baseline_missing)"
        if "Nothing to analyze" in note or exec_status == "Not Started / N/A":
            return "N/A or API unavailable — see Evidence Note"
        if has_paths:
            return "OK — file:line evidence linked in report"
        return "Metric calculated; see Evidence Note"

    # Evidence not available but tool ran
    if "All Uses" in l3 or "All Definition" in l3 or "Data Flow" in l3:
        return (
            "Beniget/def-use runner scanned Python sandbox scripts (defs_uses.json), "
            "not Java AST — gate cannot link Java file:line evidence"
        )
    if "Path Coverage" in l3:
        return (
            "JaCoCo XML present but reports line/branch counters only — "
            "workbook Path Coverage L4 metrics need path-level evidence JaCoCo does not emit"
        )
    if "Mutation" in l3 or tool == "PIT":
        if not artifact_index.get("PIT"):
            return "PIT output file not found in QA folder — mutation metrics scored without PIT XML evidence linkage"
        return "PIT ran but gate did not map mutation survivors to file:line for this L4 classification"
    if "Coverage Delta" in l3:
        return (
            "coverage_delta JSON has current_total=0 baseline_total=0 — "
            "delta engine could not pair JaCoCo current + baseline on worker"
        )
    if "OWASP" in tool or "Dependency-Check" in tool:
        if artifact_index.get("OWASP"):
            return (
                "OWASP XML present (692KB) — aggregate SCA score computed but "
                "sub-classifications (license/transitive/outdated) lack per-finding evidence paths in gate"
            )
        return "OWASP output missing"
    if "SpotBugs" in tool:
        if not artifact_index.get("SpotBugs"):
            return "SpotBugs XML not exported to QA folder — SAST sub-metrics scored without drill-down evidence"
        return "SpotBugs output present but no evidence paths for this security sub-classification"
    if "CK + PMD" in tool or "Cyclomatic" in l3:
        return (
            "CK class.csv + PMD XML present — cyclomatic aggregate scored but "
            "specific L4 labels (Decision Coverage, Test Prioritization) have no evidence mapper in gate"
        )
    if "Gitleaks" in tool or "Secret Detection" in l3:
        if artifact_index.get("Gitleaks"):
            return (
                "gitleaks.json HAS AuthService.java:23 finding — "
                "Security sheet L4 'Hardcoded Secret Detection' lacks evidence linker; "
                "Compliance sheet 'Hardcoded Secret Scan' DOES show evidence (same tool, different L4 mapping)"
            )
        return "Secret scanner output missing"
    if "git log" in tool.lower() or "pydriller" in tool.lower() or "Churn" in row.get("L4 Classification", ""):
        if artifact_index.get("Git churn"):
            return (
                "git-churn.json present with Java file stats — "
                "Performance 'Code Churn in Performance-Critical Paths' has no path filter evidence"
            )
        return "Git churn output missing or pydriller skipped"
    if "Privileged" in row.get("L4 Classification", ""):
        return "access_control.json skipped (not_github_or_gitlab) — live GitHub token unavailable on worker"
    return (
        "Tool likely ran and produced a score, but taxonomy gate has no L4→evidence-path mapper "
        "for this classification (score-only metric)"
    )


def build_artifact_index(folder: Path) -> dict[str, bool]:
    idx: dict[str, bool] = {}
    for p in folder.iterdir() if folder.is_dir() else []:
        kind = classify_artifact(p)
        idx[kind] = True
        if "PMD" in kind:
            idx["PMD"] = True
        if "JaCoCo" in kind:
            idx["JaCoCo"] = True
        if "Gitleaks" in kind:
            idx["Gitleaks"] = True
        if "Git churn" in kind:
            idx["Git churn"] = True
        if "OWASP" in kind:
            idx["OWASP"] = True
    return idx


def list_qa_files(folder: Path) -> list[dict]:
    rows = []
    for p in sorted(folder.iterdir()) if folder.is_dir() else []:
        if not p.is_file():
            continue
        status, detail = artifact_quality(p)
        rows.append(
            {
                "File Name": p.name,
                "Size (bytes)": p.stat().st_size,
                "Inferred Tool": classify_artifact(p),
                "Usability": status,
                "Notes": detail,
                "Java-Relevant": "Yes"
                if status == "Valid"
                and classify_artifact(p)
                not in ("Other / non-Java",)
                and p.name not in ARTIFACT_NOTES
                else ("Partial" if p.name in ("coverage_delta (2).json", "pmd (3).xml") else "No"),
            }
        )
    return rows


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws, max_width=70):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = width


def write_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    style_header(ws)
    autosize(ws)


def main() -> int:
    if not REPORT_HTML.is_file():
        print(f"Missing report: {REPORT_HTML}", file=sys.stderr)
        return 1
    if not TOOL_DIR.is_dir():
        print(f"Missing tool dir: {TOOL_DIR}", file=sys.stderr)
        return 1

    summary, metric_rows = parse_report(REPORT_HTML)
    tool_summary = build_tool_summary(metric_rows)
    artifact_index = build_artifact_index(TOOL_DIR)
    qa_files = list_qa_files(TOOL_DIR)
    expected_rows = map_expected_coverage()

    enriched = []
    for r in metric_rows:
        er = dict(r)
        er["Has File:Line Evidence"] = "Yes" if r.get("Evidence Paths (sample)") else "No"
        er["Evidence Gap Reason"] = evidence_reason(r, artifact_index)
        if "Evidence not available" in r.get("Evidence Note", ""):
            er["Issue Category"] = "Evidence linkage gap"
        elif r.get("Tool Execution Status") == "Failed":
            er["Issue Category"] = "Tool failure"
        elif r.get("Tool Execution Status") == "Not Started / N/A":
            er["Issue Category"] = "Not started / N/A"
        elif r.get("Metric Calculated") == "Yes":
            er["Issue Category"] = "OK — metric calculated"
        else:
            er["Issue Category"] = "Other"
        enriched.append(er)

    evidence_gaps = [r for r in enriched if "Evidence not available" in r.get("Evidence Note", "")]

    wb = Workbook()
    ws = wb.active
    ws.title = "Run Summary"
    ws.append(["Field", "Value"])
    for k, v in summary.items():
        ws.append([k, v])
    ws.append([])
    ws.append(["QA tool output folder", str(TOOL_DIR)])
    ws.append(["Files in folder", len(qa_files)])
    ws.append(["Total gate metrics", len(metric_rows)])
    ws.append(["Evidence not available count", len(evidence_gaps)])
    ws.append(["Tool failures", sum(1 for r in metric_rows if r["Tool Execution Status"] == "Failed")])
    ws.append(["Not started / N/A", sum(1 for r in metric_rows if r["Tool Execution Status"] == "Not Started / N/A")])
    ws.append(["Metrics calculated (numeric)", sum(1 for r in metric_rows if r["Metric Calculated"] == "Yes")])
    ws.append(["With file:line evidence", sum(1 for r in metric_rows if r.get("Evidence Paths (sample)"))])
    style_header(ws)
    autosize(ws)

    write_sheet(wb, "Expected Java Tools", expected_rows)
    write_sheet(wb, "QA Folder Files", qa_files)
    write_sheet(wb, "Tool Summary (Gate)", tool_summary)
    write_sheet(wb, "Metrics Detail", enriched)
    write_sheet(wb, "Evidence Not Available", evidence_gaps)

    # Tool rollup for evidence gaps
    by_tool: dict[str, list] = defaultdict(list)
    for r in evidence_gaps:
        by_tool[r["Primary Tool (mapped)"]].append(r["L4 Classification"])
    gap_tool_rows = [
        {
            "Primary Tool": tool,
            "Evidence-N/A Metrics": len(items),
            "Classifications": "; ".join(items),
        }
        for tool, items in sorted(by_tool.items(), key=lambda x: -len(x[1]))
    ]
    write_sheet(wb, "Evidence Gap by Tool", gap_tool_rows)

    non_java = [r for r in qa_files if r["Java-Relevant"] == "No"]
    write_sheet(wb, "Non-Java Artifacts", non_java)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
