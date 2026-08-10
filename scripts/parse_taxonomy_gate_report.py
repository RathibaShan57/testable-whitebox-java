#!/usr/bin/env python3
"""Parse Taxonomy gate HTML report and export tool/metric Excel analysis."""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# L3 Technique -> Primary tool(s) from Testable_Strategy_Metrics_Mapping_v0.2 (Java)
TECHNIQUE_TOOL_MAP = {
    "Code Duplication": "CPD",
    "Cyclomatic Complexity": "CK + PMD",
    "Cognitive Complexity": "PMD",
    "Lint / Rule Violations": "Checkstyle + PMD",
    "Static Vulnerabilities (SAST)": "SpotBugs (+ FindSecBugs)",
    "Dependency Risk (SCA)": "OWASP Dependency-Check",
    "Statement Coverage": "JaCoCo",
    "Branch Coverage": "JaCoCo",
    "Path Coverage": "JaCoCo",
    "Mutation Testing": "PIT",
    "Data Flow Analysis": "JaCoCo + static DU",
    "Algorithmic Complexity": "Lizard + PMD",
    "Database Query Analysis": "Semgrep",
    "Memory Management": "Semgrep + PMD",
    "Concurrency Analysis": "SpotBugs",
    "Dependency Analysis": "PMD + ArchUnit + GitHub Actions",
    "Technical Debt": "git log / pydriller",
    "Test Coverage": "JaCoCo",
    "Secret Detection": "Gitleaks + detect-secrets + Trufflehog",
    "IaC Security": "Checkov + tfsec + kics",
    "IaC Scanning": "Checkov + tfsec + kics",
    "Change Management Testing": "GitHub Branch Protection API",
    "Access Control": "GitHub API",
    "Student PII Exposure Scan": "Microsoft Presidio",
    "PII Logging Detection": "Semgrep (custom rules)",
    "User Journey Confidence": "Playwright",
    "Interface Reliability": "Playwright / Newman",
    "Response Integrity Testing": "Playwright / Newman",
    "Contract Reliability": "Playwright / Newman",
    "Consumer-Driven Contract Testing": "Pact",
    "Schema Drift Detection": "Playwright / Newman",
    "Load Signal Delta": "k6",
    "DAST — Dynamic Application Security Testing": "OWASP ZAP",
}

RUNNER_TOOL_MAP = {
    "java-perf-dependency": "PMD (java-perf-dependency runner)",
    "owasp-depcheck": "OWASP Dependency-Check",
}


def parse_report(html_path: Path) -> tuple[dict, list[dict]]:
    soup = BeautifulSoup(html_path.read_text(encoding="utf-8"), "html.parser")

    summary: dict[str, str] = {}
    for tr in soup.select("table.sum tr"):
        th, td = tr.find("th"), tr.find("td")
        if th and td:
            summary[th.get_text(strip=True)] = td.get_text(strip=True)

    rows: list[dict] = []
    current_l2 = None
    current_l3 = None
    current_l4 = None
    l2_score = None

    for el in soup.body.children:
        name = getattr(el, "name", None)
        if name == "h2":
            current_l2 = el.get_text(strip=True)
            current_l3 = None
            current_l4 = None
            l2_score = None
        elif name == "p" and "meta" in (el.get("class") or []):
            m = re.search(r"(\d+)/100", el.get_text())
            if m:
                l2_score = m.group(1)
        elif name == "h3":
            current_l3 = el.get_text(strip=True)
            current_l4 = None
        elif name == "h4":
            current_l4 = el.get_text(strip=True)
        elif name == "table" and "grid" in (el.get("class") or []):
            for tr in el.select("tbody tr"):
                tds = tr.find_all("td")
                if len(tds) < 5:
                    continue

                cls = tds[0].select_one(".cls-name")
                classification = cls.get_text(strip=True) if cls else ""
                note_el = tds[0].select_one(".evidence-note")
                note = note_el.get_text(" ", strip=True) if note_el else ""
                paths = [
                    li.get_text(strip=True).split(" — ")[0]
                    for li in tds[0].select("ul.evidence li")
                ]

                val_strong = tds[1].select_one("strong")
                value = val_strong.get_text(strip=True) if val_strong else ""
                threshold_el = tds[1].select_one(".sub")
                threshold = threshold_el.get_text(strip=True) if threshold_el else ""

                exec_span = tds[2].select_one("span")
                exec_classes = " ".join(exec_span.get("class", [])) if exec_span else ""
                if "exec-tick" in exec_classes:
                    exec_status = "Success"
                elif "exec-fail" in exec_classes:
                    exec_status = "Failed"
                elif "exec-na" in exec_classes:
                    exec_status = "Not Started / N/A"
                else:
                    exec_status = exec_span.get_text(strip=True) if exec_span else "Unknown"

                rp = tds[3].select_one(".rp")
                result = rp.get_text(strip=True) if rp else ""
                coverage = tds[4].get_text(strip=True)

                runner_id = ""
                tool_failure_detail = ""
                tm = re.search(r"Tool failure — ([^:]+):\s*(.*)", note)
                if tm:
                    runner_id = tm.group(1).strip()
                    tool_failure_detail = tm.group(2).strip()

                primary_tool = RUNNER_TOOL_MAP.get(runner_id) or TECHNIQUE_TOOL_MAP.get(
                    current_l3 or "", "(see technique mapping)"
                )
                if current_l4 and current_l4 not in (current_l3 or ""):
                    technique_label = f"{current_l3} / {current_l4}"
                else:
                    technique_label = current_l3 or ""

                tool_ran_ok = exec_status == "Success"
                has_numeric_score = bool(re.match(r"^\d+/100$", value))
                metric_calculated = tool_ran_ok and has_numeric_score
                # Calculated correctly = tool ran and produced a normalized score (PASS/WARN/FAIL all valid)
                calculation_correct = metric_calculated and result in ("PASS", "WARN", "FAIL")
                threshold_met = result == "PASS"

                rows.append(
                    {
                        "L2 Testing Type": current_l2,
                        "L2 Score (/100)": l2_score,
                        "L3 Technique": technique_label,
                        "L4 Classification": classification,
                        "Primary Tool (mapped)": primary_tool,
                        "Runner ID": runner_id or "—",
                        "Metric Value": value,
                        "Threshold": threshold,
                        "Coverage %": coverage,
                        "Tool Ran Successfully": "Yes" if tool_ran_ok else "No",
                        "Metric Calculated": "Yes" if metric_calculated else "No",
                        "Calculation Correct": "Yes" if calculation_correct else "No",
                        "Threshold Met (PASS)": "Yes" if threshold_met else "No",
                        "Result": result,
                        "Tool Execution Status": exec_status,
                        "Evidence Paths (sample)": "; ".join(paths[:3]),
                        "Evidence Note": note[:300],
                        "Tool Failure Detail": tool_failure_detail[:300],
                    }
                )

    return summary, rows


def build_tool_summary(rows: list[dict]) -> list[dict]:
    by_tool: dict[str, dict] = defaultdict(
        lambda: {
            "Primary Tool": "",
            "Metrics Total": 0,
            "Tool Ran OK": 0,
            "Tool Failed": 0,
            "Not Started / N/A": 0,
            "Metric Calculated": 0,
            "Calculation Correct": 0,
            "Threshold PASS": 0,
            "Threshold WARN": 0,
            "Threshold FAIL": 0,
            "Runner Failures": set(),
            "L3 Techniques": set(),
        }
    )

    for r in rows:
        tool = r["Primary Tool (mapped)"]
        bucket = by_tool[tool]
        bucket["Primary Tool"] = tool
        bucket["Metrics Total"] += 1
        bucket["L3 Techniques"].add(r["L3 Technique"])

        if r["Tool Ran Successfully"] == "Yes":
            bucket["Tool Ran OK"] += 1
        elif r["Tool Execution Status"] == "Failed":
            bucket["Tool Failed"] += 1
        else:
            bucket["Not Started / N/A"] += 1

        if r["Metric Calculated"] == "Yes":
            bucket["Metric Calculated"] += 1
        if r["Calculation Correct"] == "Yes":
            bucket["Calculation Correct"] += 1
        if r["Result"] == "PASS":
            bucket["Threshold PASS"] += 1
        elif r["Result"] == "WARN":
            bucket["Threshold WARN"] += 1
        elif r["Result"] == "FAIL":
            bucket["Threshold FAIL"] += 1
        if r["Runner ID"] != "—":
            bucket["Runner Failures"].add(r["Runner ID"])

    out = []
    for tool, b in sorted(by_tool.items(), key=lambda x: x[0].lower()):
        total = b["Metrics Total"]
        ran_ok = b["Tool Ran OK"]
        overall_tool_status = (
            "SUCCESS"
            if ran_ok == total and b["Tool Failed"] == 0
            else "PARTIAL"
            if ran_ok > 0
            else "FAILED"
            if b["Tool Failed"] > 0
            else "NOT RUN"
        )
        out.append(
            {
                "Primary Tool": tool,
                "Overall Tool Status": overall_tool_status,
                "Metrics Total": total,
                "Tool Ran Successfully (count)": ran_ok,
                "Tool Failed (count)": b["Tool Failed"],
                "Not Started / N/A (count)": b["Not Started / N/A"],
                "Metrics Calculated": b["Metric Calculated"],
                "Calculations Correct": b["Calculation Correct"],
                "Calc Success Rate %": round(100 * b["Calculation Correct"] / total, 1) if total else 0,
                "Threshold PASS": b["Threshold PASS"],
                "Threshold WARN": b["Threshold WARN"],
                "Threshold FAIL": b["Threshold FAIL"],
                "Runner Failures": ", ".join(sorted(b["Runner Failures"])) or "—",
                "L3 Techniques Covered": "; ".join(sorted(b["L3 Techniques"]))[:200],
            }
        )
    return out


def autosize(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = width


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_excel(summary: dict, rows: list[dict], tool_rows: list[dict], out_path: Path) -> None:
    wb = Workbook()

    # --- Run Summary ---
    ws = wb.active
    ws.title = "Run Summary"
    ws.append(["Field", "Value"])
    for k, v in summary.items():
        ws.append([k, v])
    ws.append([])
    ws.append(["Analysis totals", ""])
    ws.append(["Total metrics in report", len(rows)])
    ws.append(["Tool ran successfully", sum(1 for r in rows if r["Tool Ran Successfully"] == "Yes")])
    ws.append(["Tool failed", sum(1 for r in rows if r["Tool Execution Status"] == "Failed")])
    ws.append(["Not started / N/A", sum(1 for r in rows if r["Tool Execution Status"] == "Not Started / N/A")])
    ws.append(["Metrics calculated (numeric score)", sum(1 for r in rows if r["Metric Calculated"] == "Yes")])
    ws.append(["Calculations correct", sum(1 for r in rows if r["Calculation Correct"] == "Yes")])
    ws.append(["Threshold PASS", sum(1 for r in rows if r["Threshold Met (PASS)"] == "Yes")])
    ws.append(["Threshold WARN", sum(1 for r in rows if r["Result"] == "WARN")])
    ws.append(["Threshold FAIL", sum(1 for r in rows if r["Result"] == "FAIL")])
    style_header(ws)
    autosize(ws)

    # --- Tool Summary ---
    ws2 = wb.create_sheet("Tool Summary")
    if tool_rows:
        headers = list(tool_rows[0].keys())
        ws2.append(headers)
        for tr in tool_rows:
            ws2.append([tr[h] for h in headers])
    style_header(ws2)
    autosize(ws2)

    # --- Metrics Detail ---
    ws3 = wb.create_sheet("Metrics Detail")
    if rows:
        headers = list(rows[0].keys())
        ws3.append(headers)
        for r in rows:
            ws3.append([r[h] for h in headers])
    style_header(ws3)
    autosize(ws3, max_width=48)

    # --- Failed / Not Calculated ---
    ws4 = wb.create_sheet("Issues")
    issue_headers = [
        "L2 Testing Type",
        "L3 Technique",
        "L4 Classification",
        "Primary Tool",
        "Tool Execution Status",
        "Metric Value",
        "Result",
        "Issue Type",
        "Evidence Note",
    ]
    ws4.append(issue_headers)
    for r in rows:
        if r["Tool Execution Status"] == "Failed":
            issue = "Tool failed — no metric produced"
        elif r["Tool Execution Status"] == "Not Started / N/A":
            issue = "Not started or not applicable"
        elif r["Metric Calculated"] == "No":
            issue = "Tool ran but metric not calculated"
        elif r["Calculation Correct"] == "No":
            issue = "Calculation incomplete"
        else:
            continue
        ws4.append(
            [
                r["L2 Testing Type"],
                r["L3 Technique"],
                r["L4 Classification"],
                r["Primary Tool (mapped)"],
                r["Tool Execution Status"],
                r["Metric Value"],
                r["Result"],
                issue,
                r["Evidence Note"],
            ]
        )
    style_header(ws4)
    autosize(ws4, max_width=48)

    wb.save(out_path)


def main() -> int:
    html_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"c:\Users\Rathiba\Downloads\taxonomy-gate-70d4c502-217e-4b72-9936-707f6553b4ea (1).html"
    )
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else html_path.with_suffix(".analysis.xlsx")

    summary, rows = parse_report(html_path)
    tool_rows = build_tool_summary(rows)
    write_excel(summary, rows, tool_rows, out_path)
    print(f"Wrote {out_path}")
    print(f"Metrics: {len(rows)} | Tools: {len(tool_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
